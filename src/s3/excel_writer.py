from openpyxl import load_workbook
from backend_logs import get_logger

logger = get_logger("excel_writer.py")

# async def write_results_to_excel(file_path: str, all_results: dict):

#     print("all_result written excel input: ",all_results)
#     try:
#         wb = load_workbook(file_path)
#         ws = wb.active

#         # Normalize keys
#         normalized_results = {
#             k.strip().lower(): v for k, v in all_results.items()
#         }

#         for row in ws.iter_rows(min_row=2):
#             question_cell = row[1]  # Column B
#             question_text = str(question_cell.value).strip().lower() if question_cell.value else None

#             if question_text and question_text in normalized_results:
#                 result = normalized_results[question_text]
#                 page_num = result.get("page_number")

#                 # Answer
#                 row[2].value = result.get("answer")

#                 # PAGE NUMBER FIX (List Support)
#                 if isinstance(page_num, list):
#                     # remove duplicates + sort
#                     page_num = sorted(set(page_num))
#                     row[3].value = ", ".join(map(str, page_num))
#                 elif isinstance(page_num, int):
#                     row[3].value = page_num
#                 elif isinstance(page_num, str):
#                     row[3].value = page_num
#                 else:
#                     row[3].value = "N/A"

#                 # Status
#                 # row[4].value = result.get("status")

#             else:
#                 row[2].value = "Not Found"
#                 row[3].value = "N/A"
#                 # row[4].value = "NotFound"

#         wb.save(file_path)
#         logger.info(f"Excel Written Successfully!")
#         return file_path

#     except Exception as E:
#         logger.error(f"Excel Write Failed: {str(E)}")



from openpyxl import load_workbook
from backend_logs import get_logger

logger = get_logger("excel_writer.py")

async def write_results_to_excel(file_path: str, all_results: dict):
    # print("all_result written excel input: ",all_results)

    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # 1. Prepare data: Flatten the list of results per key
        # We consolidate multiple answers/pages for the same query key
        processed_data = {}
        for query_key, matches in all_results.items():
            if not matches:
                continue
            
            # Combine all unique answers and page numbers from the list of results
            answers = []
            pages = set()
            
            for m in matches:
                if m.get("answer"):
                    answers.append(f"[{m.get('found_anchor', 'Result')}]: {m.get('answer')}")
                
                # Handle page numbers (can be int, list, or str)
                pg = m.get("page_number")
                if isinstance(pg, list):
                    pages.update(pg)
                elif pg is not None:
                    pages.add(pg)

            processed_data[query_key.strip().lower()] = {
                "answer": "\n---\n".join(answers),
                "pages": sorted(list(pages))
            }

        # 2. Iterate through Excel rows
        for row in ws.iter_rows(min_row=2):
            question_cell = row[1]  # Column B
            if not question_cell.value:
                continue
                
            cell_val = str(question_cell.value).strip().lower()
            
            # 3. Flexible Matching
            # Check if cell value is in our keys OR if our keys contain the cell value
            match_key = None
            for key in processed_data.keys():
                if cell_val == key or cell_val in key:
                    match_key = key
                    break
            
            if match_key:
                data = processed_data[match_key]
                
                # Write Answer (Column C)
                row[2].value = data["answer"]
                
                # Write Page Numbers (Column D)
                if data["pages"]:
                    row[3].value = ", ".join(map(str, data["pages"]))
                else:
                    row[3].value = "N/A"
            else:
                # Only mark as Not Found if the cell was actually populated
                if row[2].value is None:
                    row[2].value = "Not Found"
                    row[3].value = "N/A"

        wb.save(file_path)
        logger.info(f"Excel Written Successfully with consolidated results!")
        return file_path

    except Exception as E:
        logger.error(f"Excel Write Failed: {str(E)}")
        raise E
    

